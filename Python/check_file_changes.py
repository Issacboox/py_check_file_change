from pathlib import Path
import shutil
import json
import re
from openpyxl import load_workbook


SUBFOLDER_PATTERN = re.compile(r"^\[(\d+)\]")
EXCEL_FILE_PATTERN = re.compile(r"\[\d+\].*\.xlsx$", re.IGNORECASE)


def row_contains_text(row, text):
    """Check if any cell in a row contains specific text."""
    return any(
        cell.value is not None and text in str(cell.value)
        for cell in row
    )


def row_has_data(row):
    """Check if a row has any real data."""
    return any(cell.value not in (None, "") for cell in row)


def has_added_data(excel_path: Path) -> bool:
    wb = None

    try:
        wb = load_workbook(excel_path, data_only=True, read_only=True)
        ws = wb.active

        start_checking = False

        for row_index, row in enumerate(ws.iter_rows(), start=1):
            if row_index <= 5:
                continue

            if not start_checking:
                if row_contains_text(row, "ตัวอย่าง"):
                    continue

                start_checking = True

            if start_checking and row_has_data(row):
                return True

        return False

    except Exception as e:
        print(f"Cannot read Excel file: {excel_path} | Error: {e}")
        return False

    finally:
        if wb is not None:
            wb.close()


def copy_excel_files(source_folder: Path, output_folder: Path):
    copied_files = []

    if not source_folder.exists():
        raise FileNotFoundError(f"Source folder not found: {source_folder}")

    output_folder.mkdir(parents=True, exist_ok=True)

    for number_folder in source_folder.iterdir():
        if not number_folder.is_dir():
            continue

        # First-level folder must be number-only, e.g. 5607, 10811
        if not number_folder.name.isdigit():
            continue

        # Find subfolders like [xxxx] folder name
        for subfolder in number_folder.rglob("*"):
            if not subfolder.is_dir():
                continue

            match = SUBFOLDER_PATTERN.match(subfolder.name)
            if not match:
                continue

            folder_code = match.group(1)

            target_folder = output_folder / folder_code
            target_folder.mkdir(parents=True, exist_ok=True)

            # Find matching Excel files inside this subfolder only
            for file_path in subfolder.iterdir():
                if not file_path.is_file():
                    continue

                if file_path.suffix.lower() != ".xlsx":
                    continue

                if not EXCEL_FILE_PATTERN.search(file_path.name):
                    continue

                copied_path = target_folder / file_path.name

                try:
                    shutil.copy2(file_path, copied_path)

                    copied_files.append({
                        "original_file_path": str(file_path),
                        "copied_file_path": str(copied_path),
                        "extracted_folder_code": folder_code,
                    })

                except Exception as e:
                    print(f"Cannot copy file: {file_path} | Error: {e}")

    return copied_files


def copy_changed_file(copied_file: Path, changes_folder: Path, folder_code: str) -> Path:
    """
    Copy changed Excel file into changes folder.

    Changes structure:
    results/changes/xxxx/[xxxx]filename.xlsx
    """
    changed_target_folder = changes_folder / folder_code
    changed_target_folder.mkdir(parents=True, exist_ok=True)

    changed_file_path = changed_target_folder / copied_file.name
    shutil.copy2(copied_file, changed_file_path)

    return changed_file_path


def main(source_folder: Path, output_folder: Path, changes_folder: Path):
    copied_files = copy_excel_files(source_folder, output_folder)

    result = []

    changes_folder.mkdir(parents=True, exist_ok=True)

    for item in copied_files:
        copied_file = Path(item["copied_file_path"])
        folder_code = item["extracted_folder_code"]

        added_data_found = has_added_data(copied_file)

        if added_data_found:
            try:
                changed_file_path = copy_changed_file(
                    copied_file=copied_file,
                    changes_folder=changes_folder,
                    folder_code=folder_code,
                )

                result.append({
                    "original_file_path": item["original_file_path"],
                    "copied_file_path": item["copied_file_path"],
                    "changed_file_path": str(changed_file_path),
                    "extracted_folder_code": folder_code,
                    "added_data_found": added_data_found,
                })

            except Exception as e:
                print(f"Cannot copy changed file: {copied_file} | Error: {e}")

    return result


if __name__ == "__main__":
    base_results = Path(__file__).resolve().parent.parent / "results"

    source = base_results / "input"
    output = base_results / "output"
    changes = base_results / "changes"

    result = main(source, output, changes)

    print(json.dumps(result, ensure_ascii=False, indent=2))